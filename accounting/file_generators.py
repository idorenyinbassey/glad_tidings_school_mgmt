"""
File generation utilities for financial reports
"""
from django.http import HttpResponse, FileResponse
from django.template.loader import render_to_string
from django.conf import settings
from io import BytesIO
import csv
import json
from decimal import Decimal

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportFileGenerator:
    """Generate downloadable files for financial reports"""
    
    def __init__(self, report_data, report_type, format_type):
        self.report_data = report_data
        self.report_type = report_type
        self.format_type = format_type.upper()
    
    def generate(self):
        """Generate file based on format type"""
        if self.format_type == 'PDF':
            return self.generate_pdf()
        elif self.format_type == 'EXCEL':
            return self.generate_excel()
        elif self.format_type == 'CSV':
            return self.generate_csv()
        else:
            raise ValueError(f"Unsupported format: {self.format_type}")
    
    def generate_pdf(self):
        """Generate PDF file"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab is required for PDF generation. Install with: pip install reportlab")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center alignment
        )
        
        title = f"{self.report_data.get('report_type', '').replace('_', ' ').title()}"
        period = self.report_data.get('period_name', '')
        
        story.append(Paragraph(f"Glad Tidings School", title_style))
        story.append(Paragraph(f"{title}", styles['Heading2']))
        story.append(Paragraph(f"Period: {period}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Generate content based on report type
        if self.report_type == 'income_statement':
            story.extend(self._generate_income_statement_pdf())
        elif self.report_type == 'balance_sheet':
            story.extend(self._generate_balance_sheet_pdf())
        elif self.report_type == 'cash_flow':
            story.extend(self._generate_cash_flow_pdf())
        elif self.report_type == 'fee_collection':
            story.extend(self._generate_fee_collection_pdf())
        elif self.report_type == 'expense_report':
            story.extend(self._generate_expense_report_pdf())
        
        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"Generated on: {self.report_data.get('generated_at', '')}", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def _generate_income_statement_pdf(self):
        """Generate income statement PDF content"""
        story = []
        styles = getSampleStyleSheet()
        
        # Revenue section
        revenue_data = [
            ['Revenue', ''],
            ['Tuition Fees', f"₦{self.report_data['total_revenue']:,.2f}"],
            ['Total Revenue', f"₦{self.report_data['total_revenue']:,.2f}"],
        ]
        
        # Expenses section
        expense_data = [
            ['Expenses', ''],
        ]
        
        for category, amount in self.report_data['expense_categories'].items():
            expense_data.append([category.title(), f"₦{amount:,.2f}"])
        
        expense_data.append(['Total Expenses', f"₦{self.report_data['total_expenses']:,.2f}"])
        
        # Net income
        net_income_data = [
            ['Net Income', f"₦{self.report_data['net_income']:,.2f}"]
        ]
        
        # Combine all data
        table_data = revenue_data + [['', '']] + expense_data + [['', '']] + net_income_data
        
        table = Table(table_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, len(revenue_data)), (-1, len(revenue_data)), colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        
        story.append(table)
        return story
    
    def _generate_balance_sheet_pdf(self):
        """Generate balance sheet PDF content"""
        story = []
        styles = getSampleStyleSheet()
        
        # Use default values if data is not available
        assets = self.report_data.get('assets', {'current_assets': 0, 'fixed_assets': 0})
        liabilities = self.report_data.get('liabilities', {'current_liabilities': 0, 'long_term_liabilities': 0})
        equity = self.report_data.get('equity', {'total_equity': 0})
        
        table_data = [
            ['Assets', ''],
            ['Current Assets', f"₦{assets.get('current_assets', 0):,.2f}"],
            ['Fixed Assets', f"₦{assets.get('fixed_assets', 0):,.2f}"],
            ['Total Assets', f"₦{(assets.get('current_assets', 0) + assets.get('fixed_assets', 0)):,.2f}"],
            ['', ''],
            ['Liabilities & Equity', ''],
            ['Current Liabilities', f"₦{liabilities.get('current_liabilities', 0):,.2f}"],
            ['Long-term Liabilities', f"₦{liabilities.get('long_term_liabilities', 0):,.2f}"],
            ['Total Liabilities', f"₦{(liabilities.get('current_liabilities', 0) + liabilities.get('long_term_liabilities', 0)):,.2f}"],
            ['Total Equity', f"₦{equity.get('total_equity', 0):,.2f}"],
        ]
        
        table = Table(table_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, 5), (-1, 5), colors.grey),
        ]))
        
        story.append(table)
        return story
    
    def _generate_cash_flow_pdf(self):
        """Generate cash flow PDF content"""
        story = []
        styles = getSampleStyleSheet()
        
        # Use default values if data is not available
        operating = self.report_data.get('operating_activities', 0)
        investing = self.report_data.get('investing_activities', 0)
        financing = self.report_data.get('financing_activities', 0)
        
        table_data = [
            ['Cash Flow Statement', ''],
            ['Operating Activities', f"₦{operating:,.2f}"],
            ['Investing Activities', f"₦{investing:,.2f}"],
            ['Financing Activities', f"₦{financing:,.2f}"],
            ['', ''],
            ['Net Cash Flow', f"₦{(operating + investing + financing):,.2f}"],
        ]
        
        table = Table(table_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        
        story.append(table)
        return story
    
    def _generate_fee_collection_pdf(self):
        """Generate fee collection report PDF content"""
        story = []
        styles = getSampleStyleSheet()
        
        # Summary data
        total_fees = float(self.report_data.get('total_fees', 0))
        total_collected = float(self.report_data.get('total_collected', 0))
        total_outstanding = float(self.report_data.get('total_outstanding', 0))
        collection_rate = self.report_data.get('collection_rate', 0)
        
        summary_data = [
            ['Fee Collection Summary', ''],
            ['Total Fees Due', f"₦{total_fees:,.2f}"],
            ['Total Collected', f"₦{total_collected:,.2f}"],
            ['Outstanding Amount', f"₦{total_outstanding:,.2f}"],
            ['Collection Rate', f"{collection_rate:.1f}%"],
        ]
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ]))
        
        story.append(table)
        return story
    
    def _generate_expense_report_pdf(self):
        """Generate expense report PDF content"""
        story = []
        styles = getSampleStyleSheet()
        
        # Summary data
        total_expenses = float(self.report_data.get('total_expenses', 0))
        expense_categories = self.report_data.get('expense_categories', {})
        
        summary_data = [
            ['Expense Report Summary', ''],
            ['Total Expenses', f"₦{total_expenses:,.2f}"],
            ['', ''],
            ['Expenses by Category', ''],
        ]
        
        # Add category breakdown
        for category, amount in expense_categories.items():
            summary_data.append([category.title(), f"₦{float(amount):,.2f}"])
        
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('BACKGROUND', (0, 3), (-1, 3), colors.grey),
        ]))
        
        story.append(table)
        return story
        
        table_data = [
            ['Assets', ''],
            ['Cash on Hand', f"₦{assets['cash_on_hand']:,.2f}"],
            ['Accounts Receivable', f"₦{assets['accounts_receivable']:,.2f}"],
            ['Total Assets', f"₦{assets['total_assets']:,.2f}"],
            ['', ''],
            ['Liabilities', ''],
            ['Unpaid Payroll', f"₦{liabilities['unpaid_payroll']:,.2f}"],
            ['Total Liabilities', f"₦{liabilities['total_liabilities']:,.2f}"],
            ['', ''],
            ['Equity', ''],
            ['Net Equity', f"₦{equity['net_equity']:,.2f}"],
        ]
        
        table = Table(table_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (0, 0), colors.grey),
            ('BACKGROUND', (0, 5), (0, 5), colors.grey),
            ('BACKGROUND', (0, 9), (0, 9), colors.grey),
        ]))
        
        story.append(table)
        return story
    
    def generate_excel(self):
        """Generate Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"
        
        # Header
        ws['A1'] = "Glad Tidings School"
        ws['A2'] = f"{self.report_data.get('report_type', '').replace('_', ' ').title()}"
        ws['A3'] = f"Period: {self.report_data.get('period_name', '')}"
        
        # Style header
        header_font = Font(bold=True, size=14)
        ws['A1'].font = header_font
        ws['A2'].font = Font(bold=True, size=12)
        
        row = 5  # Start data from row 5
        
        if self.report_type == 'income_statement':
            row = self._generate_income_statement_excel(ws, row)
        elif self.report_type == 'balance_sheet':
            row = self._generate_balance_sheet_excel(ws, row)
        elif self.report_type == 'cash_flow':
            row = self._generate_cash_flow_excel(ws, row)
        elif self.report_type == 'fee_collection':
            row = self._generate_fee_collection_excel(ws, row)
        elif self.report_type == 'expense_report':
            row = self._generate_expense_report_excel(ws, row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _generate_income_statement_excel(self, ws, start_row):
        """Generate income statement Excel content"""
        row = start_row
        
        # Revenue section
        ws[f'A{row}'] = "Revenue"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Tuition Fees"
        ws[f'B{row}'] = float(self.report_data['total_revenue'])
        row += 1
        
        ws[f'A{row}'] = "Total Revenue"
        ws[f'B{row}'] = float(self.report_data['total_revenue'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Expenses section
        ws[f'A{row}'] = "Expenses"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for category, amount in self.report_data['expense_categories'].items():
            ws[f'A{row}'] = category.title()
            ws[f'B{row}'] = float(amount)
            row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'B{row}'] = float(self.report_data['total_expenses'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Net Income
        ws[f'A{row}'] = "Net Income"
        ws[f'B{row}'] = float(self.report_data['net_income'])
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        return row + 1
    
    def _generate_balance_sheet_excel(self, ws, start_row):
        """Generate balance sheet Excel content"""
        row = start_row
        
        assets = self.report_data.get('assets', {'current_assets': 0, 'fixed_assets': 0})
        liabilities = self.report_data.get('liabilities', {'current_liabilities': 0, 'long_term_liabilities': 0})
        equity = self.report_data.get('equity', {'total_equity': 0})
        
        # Assets section
        ws[f'A{row}'] = "Assets"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Current Assets"
        ws[f'B{row}'] = float(assets.get('current_assets', 0))
        row += 1
        
        ws[f'A{row}'] = "Fixed Assets"
        ws[f'B{row}'] = float(assets.get('fixed_assets', 0))
        row += 1
        
        ws[f'A{row}'] = "Total Assets"
        ws[f'B{row}'] = float(assets.get('current_assets', 0) + assets.get('fixed_assets', 0))
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Liabilities section
        ws[f'A{row}'] = "Liabilities & Equity"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Current Liabilities"
        ws[f'B{row}'] = float(liabilities.get('current_liabilities', 0))
        row += 1
        
        ws[f'A{row}'] = "Long-term Liabilities"
        ws[f'B{row}'] = float(liabilities.get('long_term_liabilities', 0))
        row += 1
        
        ws[f'A{row}'] = "Total Liabilities"
        ws[f'B{row}'] = float(liabilities.get('current_liabilities', 0) + liabilities.get('long_term_liabilities', 0))
        row += 1
        
        ws[f'A{row}'] = "Total Equity"
        ws[f'B{row}'] = float(equity.get('total_equity', 0))
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        return row + 1
    
    def _generate_cash_flow_excel(self, ws, start_row):
        """Generate cash flow Excel content"""
        row = start_row
        
        operating = self.report_data.get('operating_activities', 0)
        investing = self.report_data.get('investing_activities', 0)
        financing = self.report_data.get('financing_activities', 0)
        
        ws[f'A{row}'] = "Cash Flow Statement"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Operating Activities"
        ws[f'B{row}'] = float(operating)
        row += 1
        
        ws[f'A{row}'] = "Investing Activities"
        ws[f'B{row}'] = float(investing)
        row += 1
        
        ws[f'A{row}'] = "Financing Activities"
        ws[f'B{row}'] = float(financing)
        row += 1
        
        ws[f'A{row}'] = "Net Cash Flow"
        ws[f'B{row}'] = float(operating + investing + financing)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        return row + 1
    
    def _generate_fee_collection_excel(self, ws, start_row):
        """Generate fee collection Excel content"""
        row = start_row
        
        total_fees = float(self.report_data.get('total_fees', 0))
        total_collected = float(self.report_data.get('total_collected', 0))
        total_outstanding = float(self.report_data.get('total_outstanding', 0))
        collection_rate = self.report_data.get('collection_rate', 0)
        
        ws[f'A{row}'] = "Fee Collection Summary"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Fees Due"
        ws[f'B{row}'] = total_fees
        row += 1
        
        ws[f'A{row}'] = "Total Collected"
        ws[f'B{row}'] = total_collected
        row += 1
        
        ws[f'A{row}'] = "Outstanding Amount"
        ws[f'B{row}'] = total_outstanding
        row += 1
        
        ws[f'A{row}'] = "Collection Rate"
        ws[f'B{row}'] = f"{collection_rate:.1f}%"
        
        return row + 1
    
    def _generate_expense_report_excel(self, ws, start_row):
        """Generate expense report Excel content"""
        row = start_row
        
        total_expenses = float(self.report_data.get('total_expenses', 0))
        expense_categories = self.report_data.get('expense_categories', {})
        
        ws[f'A{row}'] = "Expense Report Summary"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Expenses"
        ws[f'B{row}'] = total_expenses
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        ws[f'A{row}'] = "Expenses by Category"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for category, amount in expense_categories.items():
            ws[f'A{row}'] = category.title()
            ws[f'B{row}'] = float(amount)
            row += 1
        
        return row + 1

    def generate_csv(self):
        """Generate CSV file"""
        from io import StringIO
        import io
        
        # Use StringIO for CSV writing, then encode to bytes
        string_output = StringIO()
        writer = csv.writer(string_output)
        
        # Header
        writer.writerow(['Glad Tidings School'])
        writer.writerow([f"{self.report_data.get('report_type', '').replace('_', ' ').title()}"])
        writer.writerow([f"Period: {self.report_data.get('period_name', '')}"])
        writer.writerow([])  # Empty row
        
        if self.report_type == 'income_statement':
            self._generate_income_statement_csv(writer)
        elif self.report_type == 'balance_sheet':
            self._generate_balance_sheet_csv(writer)
        elif self.report_type == 'cash_flow':
            self._generate_cash_flow_csv(writer)
        elif self.report_type == 'fee_collection':
            self._generate_fee_collection_csv(writer)
        elif self.report_type == 'expense_report':
            self._generate_expense_report_csv(writer)
        
        writer.writerow([])
        writer.writerow([f"Generated on: {self.report_data.get('generated_at', '')}"])
        
        # Convert to bytes with BOM for Excel compatibility
        csv_content = string_output.getvalue()
        output = BytesIO()
        output.write('\ufeff'.encode('utf-8'))  # BOM for Excel compatibility
        output.write(csv_content.encode('utf-8'))
        output.seek(0)
        return output
    
    def _generate_income_statement_csv(self, writer):
        """Generate income statement CSV content"""
        # Revenue
        writer.writerow(['Revenue', ''])
        writer.writerow(['Tuition Fees', float(self.report_data['total_revenue'])])
        writer.writerow(['Total Revenue', float(self.report_data['total_revenue'])])
        writer.writerow([])
        
        # Expenses
        writer.writerow(['Expenses', ''])
        for category, amount in self.report_data['expense_categories'].items():
            writer.writerow([category.title(), float(amount)])
        writer.writerow(['Total Expenses', float(self.report_data['total_expenses'])])
        writer.writerow([])
        
        # Net Income
        writer.writerow(['Net Income', float(self.report_data['net_income'])])
    
    def _generate_balance_sheet_csv(self, writer):
        """Generate balance sheet CSV content"""
        assets = self.report_data.get('assets', {'current_assets': 0, 'fixed_assets': 0})
        liabilities = self.report_data.get('liabilities', {'current_liabilities': 0, 'long_term_liabilities': 0})
        equity = self.report_data.get('equity', {'total_equity': 0})
        
        # Assets
        writer.writerow(['Assets', ''])
        writer.writerow(['Current Assets', float(assets.get('current_assets', 0))])
        writer.writerow(['Fixed Assets', float(assets.get('fixed_assets', 0))])
        writer.writerow(['Total Assets', float(assets.get('current_assets', 0) + assets.get('fixed_assets', 0))])
        writer.writerow([])
        
        # Liabilities & Equity
        writer.writerow(['Liabilities & Equity', ''])
        writer.writerow(['Current Liabilities', float(liabilities.get('current_liabilities', 0))])
        writer.writerow(['Long-term Liabilities', float(liabilities.get('long_term_liabilities', 0))])
        writer.writerow(['Total Liabilities', float(liabilities.get('current_liabilities', 0) + liabilities.get('long_term_liabilities', 0))])
        writer.writerow(['Total Equity', float(equity.get('total_equity', 0))])
    
    def _generate_cash_flow_csv(self, writer):
        """Generate cash flow CSV content"""
        operating = self.report_data.get('operating_activities', 0)
        investing = self.report_data.get('investing_activities', 0)
        financing = self.report_data.get('financing_activities', 0)
        
        writer.writerow(['Cash Flow Statement', ''])
        writer.writerow(['Operating Activities', float(operating)])
        writer.writerow(['Investing Activities', float(investing)])
        writer.writerow(['Financing Activities', float(financing)])
        writer.writerow([])
        writer.writerow(['Net Cash Flow', float(operating + investing + financing)])
    
    def _generate_fee_collection_csv(self, writer):
        """Generate fee collection CSV content"""
        total_fees = float(self.report_data.get('total_fees', 0))
        total_collected = float(self.report_data.get('total_collected', 0))
        total_outstanding = float(self.report_data.get('total_outstanding', 0))
        collection_rate = self.report_data.get('collection_rate', 0)
        
        writer.writerow(['Fee Collection Summary', ''])
        writer.writerow(['Total Fees Due', total_fees])
        writer.writerow(['Total Collected', total_collected])
        writer.writerow(['Outstanding Amount', total_outstanding])
        writer.writerow(['Collection Rate', f"{collection_rate:.1f}%"])
    
    def _generate_expense_report_csv(self, writer):
        """Generate expense report CSV content"""
        total_expenses = float(self.report_data.get('total_expenses', 0))
        expense_categories = self.report_data.get('expense_categories', {})
        
        writer.writerow(['Expense Report Summary', ''])
        writer.writerow(['Total Expenses', total_expenses])
        writer.writerow([])
        writer.writerow(['Expenses by Category', ''])
        
        for category, amount in expense_categories.items():
            writer.writerow([category.title(), float(amount)])

def generate_report_file(report_data, report_type, format_type):
    """Main function to generate report files"""
    generator = ReportFileGenerator(report_data, report_type, format_type)
    return generator.generate()
